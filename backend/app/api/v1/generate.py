"""AI conversation generation endpoints."""

import io
import uuid
from datetime import datetime, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.core.config import settings
from app.core.rate_limiter import check_daily_quota, limiter
from app.core.rate_limiter import get_client_ip, quota_manager
from app.schemas.chat import (
    BatchResponse,
    ChatMessage,
    GenerateRequest,
    GenerateResponse,
    MessageType,
    ParseExcelResponse,
    SpeakerType,
)
from app.services.ai_service import AIService
from app.services.batch_service import BatchParseError, batch_service
from app.services.excel_service import ExcelParseError, ExcelService

router = APIRouter()
ai_service = AIService()
excel_service = ExcelService()


def format_excel_error_detail(error: ExcelParseError) -> dict:
    """Format ExcelParseError into API response format."""
    detail = {
        "error": "excel_parse_error",
        "message": error.message,
    }
    if error.row is not None:
        detail["row"] = error.row
    if error.column is not None:
        detail["column"] = error.column
    return detail


@router.post(
    "/conversation",
    response_model=GenerateResponse,
    summary="Generate AI conversation",
    description="Generate a chat conversation using AI based on the provided prompt.",
)
@limiter.limit(f"{settings.rate_limit_per_minute}/minute")
async def generate_conversation(
    request: Request,
    body: GenerateRequest,
    _: Annotated[None, Depends(check_daily_quota)],
) -> GenerateResponse:
    """
    Generate a chat conversation using AI.

    - **prompt**: Description of the conversation to generate
    - **theme**: Chat theme (kakao, instagram, telegram, discord, imessage)
    - **message_count**: Number of messages to generate (2-50)
    - **style**: Conversation style (casual, formal, romantic, funny, dramatic)
    - **language**: Output language (ko, en, ja)
    - **provider**: AI provider to use (openai, upstage)
    """
    try:
        messages, metadata = await ai_service.generate_conversation(
            prompt=body.prompt,
            message_count=body.message_count,
            style=body.style,
            language=body.language,
            provider=body.provider,
        )

        return GenerateResponse(
            success=True,
            messages=messages,
            metadata=metadata,
            tokens_used=metadata.get("tokens_used"),
            provider=body.provider,
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "validation_error", "message": str(e)},
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "generation_failed",
                "message": "Failed to generate conversation. Please try again.",
            },
        )


@router.post(
    "/parse-excel",
    response_model=ParseExcelResponse,
    summary="Parse Excel file to chat messages",
    description="Upload an Excel file and convert it to chat messages.",
)
@limiter.limit(f"{settings.rate_limit_per_minute}/minute")
async def parse_excel(
    request: Request,
    file: Annotated[UploadFile, File(description="Excel file (.xlsx, .xls)")],
    _: Annotated[None, Depends(check_daily_quota)],
) -> ParseExcelResponse:
    """
    Parse an Excel file and convert it to chat messages.

    Expected Excel format:
    | speaker | text | type | time |
    |---------|------|------|------|
    | me      | Hi!  | text | 10:30|
    | other   | Hey  | text | 10:31|
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "invalid_file", "message": "No filename provided"},
        )

    allowed_extensions = {".xlsx", ".xls"}
    file_ext = "." + file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""

    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "invalid_file_type",
                "message": f"Only Excel files allowed: {', '.join(allowed_extensions)}",
            },
        )

    # Validate file size (max 5MB)
    max_size = 5 * 1024 * 1024
    content = await file.read()
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail={"error": "file_too_large", "message": "File size exceeds 5MB limit"},
        )

    try:
        messages, stats = await excel_service.parse_excel(content, file.filename)

        return ParseExcelResponse(
            success=True,
            messages=messages,
            total_rows=stats["total_rows"],
            parsed_rows=stats["parsed_rows"],
            errors=stats.get("errors", []),
            warnings=stats.get("warnings", []),
        )

    except ExcelParseError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=format_excel_error_detail(e),
        )
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "validation_error", "message": str(e)},
        )


@router.post(
    "/demo",
    response_model=GenerateResponse,
    summary="Generate demo conversation (no AI)",
    description="Generate a demo conversation without using AI API.",
)
async def generate_demo(request: Request) -> GenerateResponse:
    """Generate a demo conversation for testing purposes."""
    demo_messages = [
        ChatMessage(
            id=str(uuid.uuid4()),
            speaker=SpeakerType.OTHER,
            speaker_name="지수",
            text="오늘 저녁 뭐 먹을까?",
            type=MessageType.TEXT,
            timestamp=datetime.now(timezone.utc),
        ),
        ChatMessage(
            id=str(uuid.uuid4()),
            speaker=SpeakerType.ME,
            speaker_name="나",
            text="음... 치킨 어때?",
            type=MessageType.TEXT,
            timestamp=datetime.now(timezone.utc),
        ),
        ChatMessage(
            id=str(uuid.uuid4()),
            speaker=SpeakerType.OTHER,
            speaker_name="지수",
            text="좋아! 어디서 시킬까?",
            type=MessageType.TEXT,
            timestamp=datetime.now(timezone.utc),
        ),
        ChatMessage(
            id=str(uuid.uuid4()),
            speaker=SpeakerType.ME,
            speaker_name="나",
            text="교촌 어때? 허니콤보 최고지",
            type=MessageType.TEXT,
            timestamp=datetime.now(timezone.utc),
        ),
        ChatMessage(
            id=str(uuid.uuid4()),
            speaker=SpeakerType.OTHER,
            speaker_name="지수",
            text="ㅋㅋㅋ 완전 동의! 콜!",
            type=MessageType.TEXT,
            timestamp=datetime.now(timezone.utc),
        ),
    ]

    return GenerateResponse(
        success=True,
        messages=demo_messages,
        metadata={"source": "demo", "generated_at": datetime.now(timezone.utc).isoformat()},
        tokens_used=0,
        provider="demo",
    )


@router.get(
    "/template",
    summary="Download Excel template",
    description="Download an Excel template for bulk conversation import.",
    response_class=StreamingResponse,
)
async def download_template() -> StreamingResponse:
    """
    Download an Excel template for bulk conversation import.

    The template includes:
    - Messages sheet with example data
    - Instructions sheet with usage guide
    """
    workbook = Workbook()

    # === Messages Sheet ===
    messages_sheet = workbook.active
    messages_sheet.title = "Messages"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        ("speaker", "발신자 (필수)"),
        ("text", "메시지 내용 (필수)"),
        ("name", "표시 이름 (선택)"),
        ("type", "타입 (선택)"),
        ("time", "시간 (선택)"),
    ]

    for col, (_, header_text) in enumerate(headers, 1):
        cell = messages_sheet.cell(row=1, column=col, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Example data
    example_data = [
        ("me", "안녕! 오늘 뭐해?", "나", "text", "14:30"),
        ("other", "나 지금 집에서 쉬고 있어~", "친구", "text", "14:31"),
        ("me", "심심하면 나와! 카페 갈래?", "나", "text", "14:32"),
        ("other", "오 좋아! 어디로 갈까?", "친구", "text", "14:33"),
        ("me", "강남역 스타벅스 어때?", "나", "text", "14:34"),
        ("other", "ㅋㅋ 완전 좋아~ 30분 뒤에 볼까?", "친구", "text", "14:35"),
        ("me", "응 그래! 이따 봐~", "나", "text", "14:36"),
        ("other", "👍", "친구", "emoji", "14:36"),
    ]

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = messages_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Set column widths
    messages_sheet.column_dimensions["A"].width = 15
    messages_sheet.column_dimensions["B"].width = 40
    messages_sheet.column_dimensions["C"].width = 15
    messages_sheet.column_dimensions["D"].width = 12
    messages_sheet.column_dimensions["E"].width = 12

    # === Instructions Sheet ===
    instructions_sheet = workbook.create_sheet("사용법")

    instructions = [
        ("TalkStudio 엑셀 템플릿 사용법", True),
        ("", False),
        ("■ 필수 열", True),
        ("speaker: 발신자 구분", False),
        ("  - me, 나, 본인, 1 → 내가 보낸 메시지", False),
        ("  - other, 상대방, 친구 → 상대방 메시지", False),
        ("  - system, 시스템 → 시스템 메시지", False),
        ("", False),
        ("text: 메시지 내용 (최대 5000자)", False),
        ("", False),
        ("■ 선택 열", True),
        ("name: 채팅방에 표시될 이름", False),
        ("type: 메시지 타입 (text, emoji, image, system)", False),
        ("time: 메시지 시간 (HH:MM 또는 YYYY-MM-DD HH:MM)", False),
        ("", False),
        ("■ 제한사항", True),
        ("- 최대 파일 크기: 5MB", False),
        ("- 최대 행 수: 1,000행", False),
        ("- 지원 형식: .xlsx, .xls, .xlsm", False),
        ("", False),
        ("■ 팁", True),
        ("- 예시 데이터를 삭제하고 실제 데이터를 입력하세요", False),
        ("- 열 이름은 한글/영어 모두 인식됩니다", False),
        ("  (speaker = 발신자 = 보낸사람 = who = from)", False),
        ("  (text = 내용 = 메시지 = message = content)", False),
    ]

    for row_idx, (text, is_header) in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row_idx, column=1, value=text)
        if is_header:
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)

    instructions_sheet.column_dimensions["A"].width = 60

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"talkstudio_template_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/batch-template",
    summary="Download batch prompt template",
    description="Download an Excel template for batch conversation generation.",
    response_class=StreamingResponse,
)
async def download_batch_template() -> StreamingResponse:
    """
    Download an Excel template for batch conversation generation.

    The template includes:
    - Prompts sheet with example prompts
    - Instructions sheet with usage guide
    """
    workbook = Workbook()

    # === Prompts Sheet ===
    prompts_sheet = workbook.active
    prompts_sheet.title = "Prompts"

    # Header style
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        ("prompt", "프롬프트 (필수)\n대화 상황 설명"),
        ("message_count", "메시지 수\n(기본: 10)"),
        ("style", "스타일\n(기본: casual)"),
        ("language", "언어\n(기본: ko)"),
        ("provider", "AI 제공자\n(기본: upstage)"),
        ("theme", "테마\n(기본: kakao)"),
    ]

    for col, (_, header_text) in enumerate(headers, 1):
        cell = prompts_sheet.cell(row=1, column=col, value=header_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set row height for header
    prompts_sheet.row_dimensions[1].height = 40

    # Example data
    example_data = [
        ("친구와 카페에서 커피 마시며 주말 계획 세우기", 8, "casual", "ko", "upstage", "kakao"),
        ("직장 동료와 프로젝트 미팅 후 점심 메뉴 정하기", 10, "formal", "ko", "upstage", "discord"),
        ("연인과 영화 보고 나서 감상 나누기", 6, "romantic", "ko", "upstage", "imessage"),
        ("친구들과 여행 계획 세우기", 12, "funny", "ko", "upstage", "telegram"),
        ("가족과 명절 일정 조율하기", 8, "casual", "ko", "upstage", "kakao"),
    ]

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = prompts_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(wrap_text=True)

    # Set column widths
    prompts_sheet.column_dimensions["A"].width = 50
    prompts_sheet.column_dimensions["B"].width = 12
    prompts_sheet.column_dimensions["C"].width = 12
    prompts_sheet.column_dimensions["D"].width = 10
    prompts_sheet.column_dimensions["E"].width = 14
    prompts_sheet.column_dimensions["F"].width = 12

    # === Instructions Sheet ===
    instructions_sheet = workbook.create_sheet("사용법")

    instructions = [
        ("TalkStudio 배치 생성 템플릿 사용법", True),
        ("", False),
        ("■ 필수 열", True),
        ("prompt: 생성할 대화 상황을 자세히 설명 (10-2000자)", False),
        ("  예: '친구와 카페에서 커피 마시며 주말 여행 계획 세우기'", False),
        ("  예: '연인과 영화관에서 영화 보고 나서 감상 나누기'", False),
        ("", False),
        ("■ 선택 열 (비워두면 기본값 적용)", True),
        ("message_count: 생성할 메시지 수 (2-50, 기본: 10)", False),
        ("style: 대화 스타일", False),
        ("  - casual: 친근하고 편안한 일상 대화 (기본)", False),
        ("  - formal: 정중하고 격식 있는 대화", False),
        ("  - romantic: 로맨틱하고 애정 어린 대화", False),
        ("  - funny: 유머러스하고 재미있는 대화", False),
        ("  - dramatic: 극적이고 감정적인 대화", False),
        ("", False),
        ("language: 출력 언어", False),
        ("  - ko: 한국어 (기본)", False),
        ("  - en: English", False),
        ("  - ja: 日本語", False),
        ("", False),
        ("provider: AI 제공자", False),
        ("  - upstage: Upstage Solar (기본, 추천)", False),
        ("  - openai: OpenAI GPT-4o-mini", False),
        ("", False),
        ("theme: 채팅 테마", False),
        ("  - kakao: 카카오톡 (기본)", False),
        ("  - instagram: 인스타그램 DM", False),
        ("  - telegram: 텔레그램", False),
        ("  - discord: 디스코드", False),
        ("  - imessage: 아이메시지", False),
        ("", False),
        ("■ 제한사항", True),
        ("- 최대 배치 크기: 100개 프롬프트", False),
        ("- 최대 파일 크기: 5MB", False),
        ("- 일일 할당량: 500회/일 (배치 포함)", False),
        ("", False),
        ("■ 팁", True),
        ("- 프롬프트는 구체적으로 작성할수록 자연스러운 대화가 생성됩니다", False),
        ("- 대화 상황, 관계, 분위기를 함께 설명하세요", False),
        ("- 예시 데이터를 삭제하고 실제 데이터를 입력하세요", False),
    ]

    for row_idx, (text, is_header) in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row_idx, column=1, value=text)
        if is_header:
            cell.font = Font(bold=True, size=12)
        else:
            cell.font = Font(size=11)

    instructions_sheet.column_dimensions["A"].width = 70

    # Save to BytesIO
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"talkstudio_batch_template_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post(
    "/batch",
    response_model=BatchResponse,
    summary="Batch generate conversations",
    description="Upload an Excel file with prompts to generate multiple conversations.",
)
@limiter.limit(f"{settings.rate_limit_per_minute}/minute")
async def process_batch(
    request: Request,
    file: Annotated[UploadFile, File(description="Excel file with prompts (.xlsx)")],
) -> BatchResponse:
    """
    Process batch of prompts from Excel file.

    Expected Excel format:
    | prompt | message_count | style | language | provider | theme |
    |--------|---------------|-------|----------|----------|-------|
    | ...    | 10            | casual| ko       | upstage  | kakao |

    Returns results for each prompt with generated messages or error details.
    """
    # Validate file
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "invalid_file", "message": "파일명이 없습니다"},
        )

    # Read file content
    content = await file.read()

    try:
        # Parse prompts from Excel
        prompts = await batch_service.parse_batch_excel(content, file.filename)

        # Check quota before processing
        client_ip = get_client_ip(request)
        has_quota, remaining = quota_manager.check_quota(client_ip)

        if not has_quota:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail={
                    "error": "quota_exceeded",
                    "message": "일일 할당량이 소진되었습니다",
                    "remaining": 0,
                },
            )

        if len(prompts) > remaining:
            # Warn but continue with available quota
            pass

        # Process batch
        result = await batch_service.process_batch(prompts, client_ip)

        return result

    except BatchParseError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "batch_parse_error",
                "message": e.message,
                "row": e.row,
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={
                "error": "batch_processing_failed",
                "message": f"배치 처리 중 오류가 발생했습니다: {str(e)[:100]}",
            },
        )
