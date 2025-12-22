"""
Batch processing service for generating multiple conversations from Excel prompts.
"""

import asyncio
import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from app.core.rate_limiter import quota_manager
from app.schemas.chat import (
    BatchPromptItem,
    BatchResponse,
    BatchResultItem,
    ChatMessage,
    Theme,
)
from app.services.ai_service import AIService

logger = logging.getLogger(__name__)


class BatchParseError(Exception):
    """Custom exception for batch file parsing errors."""

    def __init__(self, message: str, row: Optional[int] = None):
        self.message = message
        self.row = row
        super().__init__(self.format_message())

    def format_message(self) -> str:
        if self.row is not None:
            return f"{self.message} (행: {self.row})"
        return self.message


class BatchService:
    """Service for batch processing multiple conversation prompts."""

    MAX_BATCH_SIZE = 100
    MAX_CONCURRENT = 3
    MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}

    COLUMN_MAPPINGS = {
        "prompt": ["prompt", "프롬프트", "상황", "대화상황", "situation"],
        "message_count": ["message_count", "메시지수", "개수", "count"],
        "style": ["style", "스타일", "말투"],
        "language": ["language", "언어", "lang"],
        "provider": ["provider", "프로바이더", "ai"],
        "theme": ["theme", "테마", "플랫폼"],
    }

    def __init__(self):
        self.ai_service = AIService()

    async def parse_batch_excel(
        self,
        file_content: bytes,
        filename: str,
    ) -> List[BatchPromptItem]:
        """
        Parse Excel file containing batch prompts.

        Args:
            file_content: Raw file bytes
            filename: Original filename

        Returns:
            List of BatchPromptItem objects

        Raises:
            BatchParseError: If parsing fails
        """
        self._validate_file(file_content, filename)
        workbook = self._load_workbook(file_content)

        try:
            sheet = workbook.active
            if sheet is None:
                raise BatchParseError("엑셀 파일에 시트가 없습니다")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise BatchParseError("엑셀 파일이 비어있습니다")

            if len(rows) > self.MAX_BATCH_SIZE + 10:  # +10 for potential header rows
                raise BatchParseError(
                    f"배치 크기가 제한({self.MAX_BATCH_SIZE}개)을 초과했습니다"
                )

            header_idx, column_map = self._find_header_row(rows)
            if column_map is None or "prompt" not in column_map:
                raise BatchParseError(
                    "필수 열 'prompt' (또는 '프롬프트')를 찾을 수 없습니다"
                )

            prompts = self._parse_prompt_rows(rows, header_idx, column_map)

            if not prompts:
                raise BatchParseError("유효한 프롬프트가 없습니다")

            if len(prompts) > self.MAX_BATCH_SIZE:
                raise BatchParseError(
                    f"프롬프트 수가 제한({self.MAX_BATCH_SIZE}개)을 초과했습니다"
                )

            return prompts

        finally:
            workbook.close()

    def _validate_file(self, content: bytes, filename: str) -> None:
        """Validate file extension and size."""
        if not filename:
            raise BatchParseError("파일명이 없습니다")

        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise BatchParseError(
                f"지원하지 않는 파일 형식입니다. 지원: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )

        if len(content) > self.MAX_FILE_SIZE:
            max_mb = self.MAX_FILE_SIZE / (1024 * 1024)
            raise BatchParseError(f"파일 크기가 {max_mb}MB를 초과했습니다")

    def _load_workbook(self, file_content: bytes):
        """Load workbook with error handling."""
        try:
            return load_workbook(
                filename=io.BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except InvalidFileException:
            raise BatchParseError("손상된 엑셀 파일입니다")
        except Exception as e:
            raise BatchParseError(f"파일을 열 수 없습니다: {type(e).__name__}")

    def _find_header_row(
        self,
        rows: List[tuple],
        max_search: int = 10,
    ) -> Tuple[int, Optional[Dict[str, int]]]:
        """Find the header row and create column mapping."""
        for row_idx, row in enumerate(rows[:max_search]):
            column_map = self._try_map_columns(row)
            if column_map and "prompt" in column_map:
                return row_idx, column_map
        return -1, None

    def _try_map_columns(self, row: tuple) -> Optional[Dict[str, int]]:
        """Try to map header row to column indices."""
        if not row:
            return None

        column_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue

            cell_str = str(cell).strip().lower()
            for field, aliases in self.COLUMN_MAPPINGS.items():
                if cell_str in aliases:
                    column_map[field] = col_idx
                    break

        return column_map if column_map else None

    def _parse_prompt_rows(
        self,
        rows: List[tuple],
        header_idx: int,
        column_map: Dict[str, int],
    ) -> List[BatchPromptItem]:
        """Parse all prompt rows."""
        prompts = []
        data_rows = rows[header_idx + 1:]

        for row_offset, row in enumerate(data_rows):
            row_num = header_idx + 2 + row_offset

            try:
                prompt_item = self._parse_single_row(row, column_map, row_num)
                if prompt_item:
                    prompts.append(prompt_item)
            except (BatchParseError, ValidationError) as e:
                logger.warning("Row %d parse error: %s", row_num, str(e))
                continue

        return prompts

    def _parse_single_row(
        self,
        row: tuple,
        column_map: Dict[str, int],
        row_num: int,
    ) -> Optional[BatchPromptItem]:
        """Parse a single row into BatchPromptItem."""
        if not row or all(cell is None for cell in row):
            return None

        def get_cell(field: str) -> Any:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        prompt_raw = get_cell("prompt")
        if prompt_raw is None or not str(prompt_raw).strip():
            return None

        # Build prompt item with defaults
        data = {
            "prompt": str(prompt_raw).strip(),
        }

        # Optional fields with type conversion
        message_count = get_cell("message_count")
        if message_count is not None:
            try:
                data["message_count"] = int(message_count)
            except (ValueError, TypeError):
                pass

        style = get_cell("style")
        if style and str(style).strip().lower() in [
            "casual", "formal", "romantic", "funny", "dramatic"
        ]:
            data["style"] = str(style).strip().lower()

        language = get_cell("language")
        if language and str(language).strip().lower() in ["ko", "en", "ja"]:
            data["language"] = str(language).strip().lower()

        provider = get_cell("provider")
        if provider and str(provider).strip().lower() in ["openai", "upstage"]:
            data["provider"] = str(provider).strip().lower()

        theme = get_cell("theme")
        if theme:
            theme_str = str(theme).strip().lower()
            try:
                data["theme"] = Theme(theme_str)
            except ValueError:
                pass

        return BatchPromptItem(**data)

    async def process_batch(
        self,
        prompts: List[BatchPromptItem],
        client_ip: str,
    ) -> BatchResponse:
        """
        Process batch of prompts with concurrency control.

        Args:
            prompts: List of prompts to process
            client_ip: Client IP for quota tracking

        Returns:
            BatchResponse with all results
        """
        # Check quota before processing
        has_quota, remaining = quota_manager.check_quota(client_ip)
        if not has_quota:
            return BatchResponse(
                success=False,
                total=len(prompts),
                processed=0,
                failed=len(prompts),
                results=[
                    BatchResultItem(
                        row_number=i + 1,
                        prompt=p.prompt[:100],
                        status="skipped",
                        error="일일 할당량이 소진되었습니다",
                    )
                    for i, p in enumerate(prompts)
                ],
                remaining_quota=0,
            )

        # Limit to remaining quota
        prompts_to_process = prompts[:remaining]
        skipped_prompts = prompts[remaining:]

        semaphore = asyncio.Semaphore(self.MAX_CONCURRENT)
        results: List[BatchResultItem] = []

        async def process_one(idx: int, item: BatchPromptItem) -> BatchResultItem:
            async with semaphore:
                row_number = idx + 1
                try:
                    messages, metadata = await self.ai_service.generate_conversation(
                        prompt=item.prompt,
                        message_count=item.message_count,
                        style=item.style,
                        language=item.language,
                        provider=item.provider,
                        use_cache=True,
                    )

                    # Increment quota
                    quota_manager.increment(client_ip)

                    return BatchResultItem(
                        row_number=row_number,
                        prompt=item.prompt[:100],
                        status="success",
                        messages=messages,
                        tokens_used=metadata.get("tokens_used", 0),
                    )

                except Exception as e:
                    logger.error("Batch item %d failed: %s", row_number, str(e))
                    return BatchResultItem(
                        row_number=row_number,
                        prompt=item.prompt[:100],
                        status="failed",
                        error=str(e)[:200],
                    )

        # Process all prompts concurrently with semaphore
        tasks = [process_one(i, p) for i, p in enumerate(prompts_to_process)]
        results = await asyncio.gather(*tasks)

        # Add skipped prompts
        for i, p in enumerate(skipped_prompts):
            results.append(
                BatchResultItem(
                    row_number=len(prompts_to_process) + i + 1,
                    prompt=p.prompt[:100],
                    status="skipped",
                    error="할당량 초과로 건너뜀",
                )
            )

        # Calculate stats
        processed = sum(1 for r in results if r.status == "success")
        failed = sum(1 for r in results if r.status == "failed")

        # Get updated quota
        _, updated_remaining = quota_manager.check_quota(client_ip)

        return BatchResponse(
            success=processed > 0,
            total=len(prompts),
            processed=processed,
            failed=failed,
            results=results,
            remaining_quota=updated_remaining,
        )


# Global instance
batch_service = BatchService()
