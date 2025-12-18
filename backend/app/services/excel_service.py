"""
Excel parsing service with robust error handling.
Returns detailed validation errors instead of 500 Internal Server Errors.
"""

import io
import uuid
import zipfile
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, Field, field_validator, ValidationError

from app.schemas.chat import ChatMessage, MessageType, SpeakerType


class ExcelParseError(Exception):
    """Custom exception for Excel parsing errors with details."""

    def __init__(self, message: str, row: int | None = None, column: str | None = None):
        self.message = message
        self.row = row
        self.column = column
        super().__init__(self.format_message())

    def format_message(self) -> str:
        parts = [self.message]
        if self.row is not None:
            parts.append(f"(행: {self.row})")
        if self.column is not None:
            parts.append(f"(열: {self.column})")
        return " ".join(parts)


class RowValidation(BaseModel):
    """Pydantic model for row validation."""

    speaker: str = Field(..., min_length=1, max_length=50)
    text: str = Field(..., min_length=1, max_length=5000)
    message_type: str | None = Field(default=None, max_length=20)
    name: str | None = Field(default=None, max_length=100)
    time: datetime | str | None = None

    @field_validator("speaker")
    @classmethod
    def validate_speaker(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("발신자는 비어있을 수 없습니다")
        return v

    @field_validator("text")
    @classmethod
    def validate_text(cls, v: str) -> str:
        v = v.strip()
        if not v:
            raise ValueError("메시지 내용은 비어있을 수 없습니다")
        return v


class ExcelService:
    """Service for parsing Excel files into chat messages with robust validation."""

    COLUMN_MAPPINGS = {
        "speaker": ["speaker", "발신자", "보낸사람", "who", "from"],
        "text": ["text", "내용", "메시지", "message", "content"],
        "type": ["type", "타입", "유형"],
        "time": ["time", "시간", "timestamp", "날짜"],
        "name": ["name", "이름", "닉네임", "nickname"],
    }

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    MAX_ROWS = 1000

    async def parse_excel(
        self,
        file_content: bytes,
        filename: str,
    ) -> tuple[list[ChatMessage], dict]:
        """
        Parse Excel file content into chat messages.

        Raises:
            ExcelParseError: With detailed error message for client-side handling

        Returns:
            Tuple of (messages, stats)
        """
        # Validate file extension
        self._validate_filename(filename)

        # Validate file size
        self._validate_file_size(file_content)

        # Load workbook with error handling
        workbook = self._load_workbook_safe(file_content)

        try:
            sheet = workbook.active
            if sheet is None:
                raise ExcelParseError("엑셀 파일에 시트가 없습니다")

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise ExcelParseError("엑셀 파일이 비어있습니다")

            # Validate row count
            if len(rows) > self.MAX_ROWS:
                raise ExcelParseError(
                    f"행 수가 제한({self.MAX_ROWS}행)을 초과했습니다. "
                    f"현재 행 수: {len(rows)}"
                )

            # Find header and map columns
            header_row_idx, column_map = self._find_header_row(rows)
            if column_map is None:
                raise ExcelParseError(
                    "필수 열을 찾을 수 없습니다. "
                    "'speaker'(또는 '발신자')와 'text'(또는 '내용') 열이 필요합니다."
                )

            # Parse all data rows
            messages, errors, warnings = self._parse_all_rows(
                rows, header_row_idx, column_map
            )

            if not messages:
                if errors:
                    # Return first error as main error
                    first_error = errors[0]
                    raise ExcelParseError(
                        first_error["error"],
                        row=first_error["row"],
                    )
                raise ExcelParseError("유효한 메시지가 없습니다")

            stats = {
                "total_rows": len(rows) - header_row_idx - 1,
                "parsed_rows": len(messages),
                "error_count": len(errors),
                "errors": errors[:10],
                "warnings": warnings,
            }

            return messages, stats

        finally:
            workbook.close()

    def _validate_filename(self, filename: str) -> None:
        """Validate file extension."""
        if not filename:
            raise ExcelParseError("파일명이 없습니다")

        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ExcelParseError(
                f"지원하지 않는 파일 형식입니다. "
                f"지원 형식: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )

    def _validate_file_size(self, content: bytes) -> None:
        """Validate file size."""
        if len(content) > self.MAX_FILE_SIZE:
            max_mb = self.MAX_FILE_SIZE / (1024 * 1024)
            raise ExcelParseError(f"파일 크기가 {max_mb}MB를 초과했습니다")

    def _load_workbook_safe(self, file_content: bytes):
        """Load workbook with proper error handling."""
        try:
            return load_workbook(
                filename=io.BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except InvalidFileException:
            raise ExcelParseError(
                "손상된 엑셀 파일입니다. 파일을 다시 저장한 후 업로드해주세요."
            )
        except zipfile.BadZipFile:
            raise ExcelParseError(
                "유효하지 않은 엑셀 파일입니다. xlsx 형식인지 확인해주세요."
            )
        except Exception as e:
            raise ExcelParseError(f"파일을 열 수 없습니다: {type(e).__name__}")

    def _find_header_row(
        self,
        rows: list[tuple],
        max_search: int = 10,
    ) -> tuple[int, dict | None]:
        """Find the header row and create column mapping."""
        for row_idx, row in enumerate(rows[:max_search]):
            column_map = self._try_map_columns(row)
            if column_map and "speaker" in column_map and "text" in column_map:
                return row_idx, column_map
        return -1, None

    def _try_map_columns(self, row: tuple) -> dict[str, int] | None:
        """Try to map header row to column indices."""
        if not row:
            return None

        column_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue

            cell_str = str(cell).strip().lower()
            for field, aliases in self.COLUMN_MAPPINGS.items():
                if cell_str in aliases:
                    column_map[field] = col_idx
                    break

        return column_map if column_map else None

    def _parse_all_rows(
        self,
        rows: list[tuple],
        header_row_idx: int,
        column_map: dict[str, int],
    ) -> tuple[list[ChatMessage], list[dict], list[str]]:
        """Parse all data rows with comprehensive error collection."""
        messages = []
        errors = []
        warnings = []

        data_rows = rows[header_row_idx + 1:]

        for row_offset, row in enumerate(data_rows):
            row_num = header_row_idx + 2 + row_offset  # Excel 1-indexed + header

            try:
                message = self._parse_row_validated(row, column_map, row_num)
                if message:
                    messages.append(message)
            except ExcelParseError as e:
                errors.append({
                    "row": row_num,
                    "column": e.column,
                    "error": e.message,
                })
            except ValidationError as e:
                # Pydantic validation error
                error_msg = self._format_pydantic_error(e)
                errors.append({
                    "row": row_num,
                    "error": error_msg,
                })
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": f"예상치 못한 오류: {type(e).__name__}",
                })

        if len(errors) > 10:
            warnings.append(f"총 {len(errors)}개의 오류가 발생했습니다. 처음 10개만 표시됩니다.")

        return messages, errors, warnings

    def _format_pydantic_error(self, error: ValidationError) -> str:
        """Format Pydantic validation error into Korean message."""
        messages = []
        for err in error.errors():
            field = err.get("loc", ["unknown"])[0]
            field_ko = {
                "speaker": "발신자",
                "text": "내용",
                "message_type": "메시지 타입",
                "name": "이름",
                "time": "시간",
            }.get(str(field), str(field))

            err_type = err.get("type", "")
            if "string_too_short" in err_type:
                messages.append(f"{field_ko} 값이 너무 짧습니다")
            elif "string_too_long" in err_type:
                messages.append(f"{field_ko} 값이 너무 깁니다")
            else:
                messages.append(f"{field_ko} 형식이 잘못되었습니다")

        return "; ".join(messages) if messages else "데이터 형식 오류"

    def _parse_row_validated(
        self,
        row: tuple,
        column_map: dict[str, int],
        row_num: int,
    ) -> ChatMessage | None:
        """Parse a single row with Pydantic validation."""
        if not row or all(cell is None for cell in row):
            return None

        # Extract raw values
        raw_data = self._extract_row_data(row, column_map, row_num)
        if raw_data is None:
            return None

        # Validate with Pydantic
        validated = RowValidation(**raw_data)

        # Convert to ChatMessage
        speaker = self._parse_speaker(validated.speaker)
        msg_type = self._parse_message_type(validated.message_type)
        timestamp = self._parse_timestamp(validated.time, row_num)

        return ChatMessage(
            id=str(uuid.uuid4()),
            speaker=speaker,
            speaker_name=validated.name,
            text=validated.text,
            type=msg_type,
            timestamp=timestamp,
        )

    def _extract_row_data(
        self,
        row: tuple,
        column_map: dict[str, int],
        row_num: int,
    ) -> dict[str, Any] | None:
        """Extract raw data from row based on column mapping."""

        def get_cell(field: str) -> Any:
            idx = column_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        speaker_raw = get_cell("speaker")
        text_raw = get_cell("text")

        # Skip completely empty rows
        if speaker_raw is None and text_raw is None:
            return None

        # Validate required fields exist
        if speaker_raw is None:
            raise ExcelParseError("발신자 값이 없습니다", row=row_num, column="speaker")

        if text_raw is None:
            raise ExcelParseError("메시지 내용이 없습니다", row=row_num, column="text")

        return {
            "speaker": str(speaker_raw) if speaker_raw is not None else "",
            "text": str(text_raw) if text_raw is not None else "",
            "message_type": str(get_cell("type")) if get_cell("type") else None,
            "name": str(get_cell("name")) if get_cell("name") else None,
            "time": get_cell("time"),
        }

    def _parse_speaker(self, speaker_str: str) -> SpeakerType:
        """Parse speaker string to SpeakerType enum."""
        normalized = speaker_str.strip().lower()
        if normalized in ("me", "나", "본인", "1"):
            return SpeakerType.ME
        elif normalized in ("system", "시스템"):
            return SpeakerType.SYSTEM
        return SpeakerType.OTHER

    def _parse_message_type(self, type_str: str | None) -> MessageType:
        """Parse message type string to MessageType enum."""
        if not type_str:
            return MessageType.TEXT

        normalized = type_str.strip().lower()
        if normalized in ("emoji", "이모지", "이모티콘"):
            return MessageType.EMOJI
        elif normalized in ("image", "이미지", "사진"):
            return MessageType.IMAGE
        elif normalized in ("system", "시스템"):
            return MessageType.SYSTEM
        return MessageType.TEXT

    def _parse_timestamp(
        self,
        time_val: datetime | str | None,
        row_num: int,
    ) -> datetime:
        """Parse timestamp with error handling."""
        if time_val is None:
            return datetime.now(timezone.utc)

        if isinstance(time_val, datetime):
            return time_val.replace(tzinfo=timezone.utc)

        # Try parsing string formats
        if isinstance(time_val, str):
            time_str = time_val.strip()
            formats = [
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
                "%Y/%m/%d %H:%M:%S",
                "%Y/%m/%d %H:%M",
                "%Y.%m.%d %H:%M:%S",
                "%Y.%m.%d %H:%M",
                "%H:%M:%S",
                "%H:%M",
            ]

            for fmt in formats:
                try:
                    parsed = datetime.strptime(time_str, fmt)
                    return parsed.replace(tzinfo=timezone.utc)
                except ValueError:
                    continue

            raise ExcelParseError(
                f"날짜 형식이 잘못되었습니다: '{time_str}'. "
                "지원 형식: YYYY-MM-DD HH:MM:SS, HH:MM 등",
                row=row_num,
                column="time",
            )

        return datetime.now(timezone.utc)
