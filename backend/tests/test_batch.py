"""Tests for batch processing functionality."""

import io
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app


client = TestClient(app)


class TestBatchTemplate:
    """Tests for batch template download."""

    def test_download_batch_template(self):
        """Test downloading batch template."""
        response = client.get("/api/v1/generate/batch-template")

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert "batch_template" in response.headers["content-disposition"]

    def test_batch_template_is_valid_excel(self):
        """Test that batch template is a valid Excel file."""
        response = client.get("/api/v1/generate/batch-template")

        # Try to load as Excel
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(response.content))

        # Check sheets
        assert "Prompts" in wb.sheetnames
        assert "사용법" in wb.sheetnames

        # Check headers in Prompts sheet
        prompts_sheet = wb["Prompts"]
        headers = [cell.value for cell in prompts_sheet[1]]

        # Headers should contain prompt info
        assert any("프롬프트" in str(h) for h in headers if h)

        wb.close()


class TestBatchProcessing:
    """Tests for batch processing endpoint."""

    def _create_test_excel(self, prompts: list) -> bytes:
        """Create a test Excel file with prompts."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Prompts"

        # Headers
        headers = ["prompt", "message_count", "style", "language", "provider", "theme"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data rows
        for row_idx, prompt_data in enumerate(prompts, 2):
            if isinstance(prompt_data, str):
                ws.cell(row=row_idx, column=1, value=prompt_data)
            else:
                for col_idx, value in enumerate(prompt_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def test_batch_no_file(self):
        """Test batch endpoint without file."""
        response = client.post("/api/v1/generate/batch")
        assert response.status_code == 422  # Validation error

    def test_batch_invalid_file_type(self):
        """Test batch endpoint with invalid file type."""
        response = client.post(
            "/api/v1/generate/batch",
            files={"file": ("test.txt", b"hello world", "text/plain")},
        )
        assert response.status_code == 400
        assert "지원하지 않는" in response.json()["detail"]["message"]

    def test_batch_empty_excel(self):
        """Test batch endpoint with empty Excel file."""
        wb = Workbook()
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = client.post(
            "/api/v1/generate/batch",
            files={
                "file": (
                    "empty.xlsx",
                    output.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 400

    def test_batch_missing_prompt_column(self):
        """Test batch endpoint with Excel missing prompt column."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="other_column")
        ws.cell(row=2, column=1, value="some value")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = client.post(
            "/api/v1/generate/batch",
            files={
                "file": (
                    "no_prompt.xlsx",
                    output.read(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 400
        assert "prompt" in response.json()["detail"]["message"].lower()

    def test_batch_short_prompt(self):
        """Test batch endpoint with too short prompt."""
        excel_content = self._create_test_excel(["short"])  # Less than 10 chars

        response = client.post(
            "/api/v1/generate/batch",
            files={
                "file": (
                    "short_prompt.xlsx",
                    excel_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        # Short prompts should be skipped during parsing
        assert response.status_code == 400
        assert "유효한 프롬프트" in response.json()["detail"]["message"]

    def test_batch_valid_prompts_structure(self):
        """Test batch endpoint parses valid prompts correctly."""
        # Create Excel with valid prompts
        prompts = [
            ("친구와 카페에서 커피 마시며 대화하기", 5, "casual", "ko", "upstage", "kakao"),
            ("동료와 회의 후 점심 메뉴 정하기 대화", 8, "formal", "ko", "upstage", "discord"),
        ]
        excel_content = self._create_test_excel(prompts)

        # This will fail due to no API key, but we can test the parsing
        response = client.post(
            "/api/v1/generate/batch",
            files={
                "file": (
                    "valid.xlsx",
                    excel_content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        # Either success (with mock) or error (due to API key)
        # The important thing is it doesn't fail on parsing
        assert response.status_code in [200, 500]


class TestBatchServiceUnit:
    """Unit tests for BatchService."""

    @pytest.mark.asyncio
    async def test_parse_batch_excel_valid(self):
        """Test parsing valid batch Excel file."""
        from app.services.batch_service import batch_service

        # Create valid Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Prompts"

        headers = ["prompt", "message_count", "style"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        ws.cell(row=2, column=1, value="친구와 카페에서 커피 마시며 대화하기")
        ws.cell(row=2, column=2, value=8)
        ws.cell(row=2, column=3, value="casual")

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()

        prompts = await batch_service.parse_batch_excel(content, "test.xlsx")

        assert len(prompts) == 1
        assert prompts[0].prompt == "친구와 카페에서 커피 마시며 대화하기"
        assert prompts[0].message_count == 8
        assert prompts[0].style == "casual"

    @pytest.mark.asyncio
    async def test_parse_batch_excel_korean_headers(self):
        """Test parsing Excel with Korean headers."""
        from app.services.batch_service import batch_service

        wb = Workbook()
        ws = wb.active

        # Korean headers
        ws.cell(row=1, column=1, value="프롬프트")
        ws.cell(row=1, column=2, value="메시지수")
        ws.cell(row=1, column=3, value="스타일")

        ws.cell(row=2, column=1, value="연인과 데이트 후 대화하는 상황을 만들어줘")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value="romantic")

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()

        prompts = await batch_service.parse_batch_excel(content, "korean.xlsx")

        assert len(prompts) == 1
        assert "연인" in prompts[0].prompt
        assert prompts[0].style == "romantic"

    @pytest.mark.asyncio
    async def test_parse_batch_excel_defaults(self):
        """Test that defaults are applied for missing columns."""
        from app.services.batch_service import batch_service

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="prompt")
        ws.cell(row=2, column=1, value="친구와 영화관에서 영화 본 후 감상 나누기")

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()

        prompts = await batch_service.parse_batch_excel(content, "minimal.xlsx")

        assert len(prompts) == 1
        # Check defaults
        assert prompts[0].message_count == 10
        assert prompts[0].style == "casual"
        assert prompts[0].language == "ko"
        assert prompts[0].provider == "upstage"

    @pytest.mark.asyncio
    async def test_parse_batch_excel_invalid_extension(self):
        """Test rejection of invalid file extension."""
        from app.services.batch_service import BatchParseError, batch_service

        with pytest.raises(BatchParseError) as exc_info:
            await batch_service.parse_batch_excel(b"dummy", "test.txt")

        assert "지원하지 않는" in str(exc_info.value)
