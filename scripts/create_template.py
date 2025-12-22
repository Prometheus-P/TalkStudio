#!/usr/bin/env python3
"""
TalkStudio - Excel Template Generator

엑셀 템플릿 파일을 생성합니다.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


def create_template():
    """Discord 대화 생성용 엑셀 템플릿 생성"""
    wb = Workbook()

    # === Sheet 1: 설정 (Configuration) ===
    ws_config = wb.active
    ws_config.title = "설정"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="5865F2", end_color="5865F2", fill_type="solid")  # Discord Blurple
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 설정 시트 헤더
    config_headers = [
        ("대화ID", 15, "대화 고유 식별자 (필수)"),
        ("채널명", 20, "Discord 채널명"),
        ("대화날짜", 15, "YYYY-MM-DD 형식"),
        ("대화시간대", 15, "예: 오후 8시경"),
        ("메시지수", 12, "생성할 메시지 수"),
        ("대화맥락", 50, "AI에게 전달할 대화 상황 설명 (필수)"),
        ("대화스타일", 15, "casual/formal/dramatic"),
        ("언어", 10, "ko/en/ja"),
    ]

    for col_idx, (header, width, _) in enumerate(config_headers, 1):
        cell = ws_config.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_config.column_dimensions[get_column_letter(col_idx)].width = width

    # 설정 시트 샘플 데이터
    sample_configs = [
        ("conv_001", "game-trade", "2025-01-15", "오후 8시경", 8,
         "온라인 게임 아이템 거래. 판매자가 '전설의 검' 아이템을 5만원에 판매하려 하고, 구매자가 가격 협상 중. 최종적으로 합의하는 대화.",
         "casual", "ko"),
        ("conv_002", "친구-잡담", "2025-01-16", "오전 10시경", 6,
         "오랜만에 연락한 친구들이 주말 계획을 세우는 대화. 영화 보러 갈지, 맛집 갈지 고민 중.",
         "casual", "ko"),
        ("conv_003", "스터디-공지", "2025-01-17", "오후 3시경", 10,
         "코딩 스터디 그룹에서 다음 주 모임 일정을 정하는 대화. 3명이 참여하며 각자 가능한 시간을 조율함.",
         "casual", "ko"),
    ]

    for row_idx, data in enumerate(sample_configs, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 6:  # 대화맥락 열은 왼쪽 정렬
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # === Sheet 2: 참여자 (Participants) ===
    ws_participants = wb.create_sheet(title="참여자")

    # 참여자 시트 헤더
    participant_headers = [
        ("대화ID", 15, "설정 시트의 대화ID와 일치"),
        ("참여자ID", 15, "대화 내 고유 ID"),
        ("닉네임", 20, "표시될 이름"),
        ("역할", 10, "me 또는 other"),
        ("아바타시드", 15, "DiceBear 아바타 시드"),
        ("성격", 40, "AI가 참고할 성격 설명"),
    ]

    for col_idx, (header, width, _) in enumerate(participant_headers, 1):
        cell = ws_participants.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_participants.column_dimensions[get_column_letter(col_idx)].width = width

    # 참여자 시트 샘플 데이터
    sample_participants = [
        # conv_001: 2명 (1:1 거래)
        ("conv_001", "seller", "판매자A_demo", "me", "seller123", "협상에 능숙한 베테랑 판매자. 친절하지만 가격은 잘 안 깎아줌."),
        ("conv_001", "buyer", "구매자1_demo", "other", "buyer456", "신중하지만 적극적인 구매자. 가격 흥정을 즐김."),
        # conv_002: 2명 (친구 대화)
        ("conv_002", "user1", "철수", "me", "cheolsu", "활발하고 재미있는 성격. 이모티콘을 많이 씀."),
        ("conv_002", "user2", "영희", "other", "younghee", "차분하고 계획적인 성격. 약속을 잘 지킴."),
        # conv_003: 3명 (그룹 대화)
        ("conv_003", "admin", "스터디장", "me", "admin01", "스터디 리더. 일정 조율을 주도함."),
        ("conv_003", "member1", "민수", "other", "minsu", "유머러스하고 느긋한 성격."),
        ("conv_003", "member2", "지현", "other", "jihyun", "꼼꼼하고 책임감 있는 성격."),
    ]

    for row_idx, data in enumerate(sample_participants, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws_participants.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 6:  # 성격 열은 왼쪽 정렬
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # === Sheet 3: 사용법 (Instructions) ===
    ws_instructions = wb.create_sheet(title="사용법")

    instructions = [
        ("TalkStudio - Excel 기반 Discord 대화 생성기", ""),
        ("", ""),
        ("사용법:", ""),
        ("1. '설정' 시트에 대화 설정을 입력하세요", ""),
        ("2. '참여자' 시트에 대화 참여자를 입력하세요", ""),
        ("3. 다음 명령어로 실행:", "UPSTAGE_API_KEY=up_xxx python scripts/generate_from_excel.py scripts/templates/discord_chat_template.xlsx"),
        ("4. 이미지 캡처:", "node scripts/capture_discord_chats.cjs"),
        ("", ""),
        ("필수 열 (설정):", "대화ID, 대화맥락"),
        ("필수 열 (참여자):", "대화ID, 참여자ID, 닉네임, 역할"),
        ("", ""),
        ("역할 값:", "me (내 메시지), other (상대 메시지)"),
        ("대화스타일:", "casual (일상), formal (격식), dramatic (극적)"),
        ("언어:", "ko (한국어), en (영어), ja (일본어)"),
        ("", ""),
        ("주의사항:", ""),
        ("- 대화ID는 설정과 참여자 시트에서 동일해야 합니다", ""),
        ("- 3명 이상 그룹 대화도 지원됩니다", ""),
        ("- AI가 생성한 대화는 실제 대화가 아닌 샘플입니다", ""),
    ]

    for row_idx, (col_a, col_b) in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=col_a)
        ws_instructions.cell(row=row_idx, column=2, value=col_b)

    # 첫 번째 행 볼드
    ws_instructions.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_instructions.column_dimensions["A"].width = 50
    ws_instructions.column_dimensions["B"].width = 80

    # 저장
    output_path = Path(__file__).parent / "templates" / "discord_chat_template.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"템플릿 생성 완료: {output_path}")
    return output_path


if __name__ == "__main__":
    create_template()
