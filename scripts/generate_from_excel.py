#!/usr/bin/env python3
"""
TalkStudio - Excel-Based Discord Chat Generator

Copyright (c) 2024-2025 TalkStudio. All Rights Reserved.
PROPRIETARY AND CONFIDENTIAL. See LICENSE file for details.

엑셀 설정 파일을 읽어 Upstage API로 Discord 스타일 대화를 생성합니다.

Usage:
    UPSTAGE_API_KEY=up_xxx python scripts/generate_from_excel.py input.xlsx
    UPSTAGE_API_KEY=up_xxx python scripts/generate_from_excel.py input.xlsx --output output.json

DISCLAIMER:
- 이 스크립트로 생성된 대화는 데모/테스트 목적으로만 사용됩니다.
- 실제 대화가 아닌 AI가 생성한 가상의 샘플 데이터입니다.
- 사기, 허위 증거 조작 등 불법적인 용도로 사용을 금지합니다.
"""

import argparse
import asyncio
import json
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Optional
import random

try:
    import httpx
except ImportError:
    print("Error: httpx 패키지가 필요합니다. 'pip install httpx'로 설치하세요.")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    print("Error: openpyxl 패키지가 필요합니다. 'pip install openpyxl'로 설치하세요.")
    sys.exit(1)


# === Configuration ===
UPSTAGE_API_KEY = os.getenv("UPSTAGE_API_KEY")
UPSTAGE_BASE_URL = "https://api.upstage.ai/v1/solar"
DEFAULT_MODEL = "solar-pro"
MAX_RETRIES = 3
RETRY_DELAY = 2.0


# === Data Classes ===

@dataclass
class Participant:
    """대화 참여자 정보"""
    conversation_id: str
    participant_id: str
    nickname: str
    role: str  # me, other, participant_N (3명 이상일 때)
    avatar_seed: Optional[str] = None
    personality: Optional[str] = None

    @property
    def avatar_url(self) -> str:
        """DiceBear 아바타 URL 생성"""
        seed = self.avatar_seed or self.nickname
        return f"https://api.dicebear.com/7.x/pixel-art/svg?seed={seed}"


@dataclass
class ConversationConfig:
    """대화 설정"""
    conversation_id: str
    context: str  # 핵심: 대화 맥락 (시스템 프롬프트로 전달)
    channel_name: str = "general"
    date: Optional[datetime] = None
    time_context: Optional[str] = None  # "오후 8시경"
    message_count: int = 8
    style: str = "casual"  # casual, formal, dramatic
    language: str = "ko"  # ko, en, ja
    participants: List[Participant] = field(default_factory=list)


@dataclass
class GeneratedMessage:
    """생성된 메시지"""
    id: int
    sender: str  # participant_id
    text: str
    time: str = ""


# === Excel Parser ===

class ExcelParser:
    """엑셀 파일 파싱 클래스 (한국어/영어 헤더 지원)"""

    REQUIRED_CONFIG_COLS = {"conversation_id", "context"}
    REQUIRED_PARTICIPANT_COLS = {"conversation_id", "participant_id", "nickname", "role"}

    # 한국어 → 영어 헤더 매핑
    HEADER_MAPPING = {
        # Configuration sheet
        "대화id": "conversation_id",
        "채널명": "channel_name",
        "대화날짜": "date",
        "대화시간대": "time_context",
        "메시지수": "message_count",
        "대화맥락": "context",
        "대화스타일": "style",
        "언어": "language",
        # Participants sheet
        "참여자id": "participant_id",
        "닉네임": "nickname",
        "역할": "role",
        "아바타시드": "avatar_seed",
        "성격": "personality",
    }

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self._validate_file()

    def _validate_file(self):
        """파일 유효성 검증"""
        if not self.filepath.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {self.filepath}")
        if self.filepath.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
            raise ValueError("지원하지 않는 파일 형식입니다. xlsx, xls, xlsm만 지원됩니다.")

    def parse(self) -> List[ConversationConfig]:
        """엑셀 파일 파싱하여 대화 설정 목록 반환"""
        try:
            workbook = load_workbook(self.filepath, read_only=True, data_only=True)
        except InvalidFileException:
            raise ValueError("손상된 엑셀 파일입니다.")

        try:
            # Configuration 시트 파싱
            config_sheet = self._find_sheet(workbook, ["설정", "Configuration", "config"])
            configs = self._parse_config_sheet(config_sheet)

            # Participants 시트 파싱
            participants_sheet = self._find_sheet(
                workbook, ["참여자", "Participants", "participants"]
            )
            participants = self._parse_participants_sheet(participants_sheet)

            # 참여자를 설정에 매핑
            self._map_participants_to_configs(configs, participants)

            return configs
        finally:
            workbook.close()

    def _find_sheet(self, workbook, sheet_names: List[str]):
        """시트 이름으로 시트 찾기"""
        for name in sheet_names:
            if name in workbook.sheetnames:
                return workbook[name]

        # 시트 이름이 없으면 순서대로 사용 (첫 번째=설정, 두 번째=참여자)
        if len(workbook.sheetnames) >= 2:
            if "config" in sheet_names[0].lower() or "설정" in sheet_names[0]:
                return workbook[workbook.sheetnames[0]]
            else:
                return workbook[workbook.sheetnames[1]]

        if len(workbook.sheetnames) >= 1:
            return workbook[workbook.sheetnames[0]]

        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_names}")

    def _normalize_header(self, header_row) -> List[str]:
        """헤더 정규화 (한국어 → 영어 변환)"""
        normalized = []
        for col in header_row:
            if col is None:
                normalized.append("")
                continue
            col_str = str(col).strip().lower().replace(" ", "_")
            # 한국어 매핑 적용
            mapped = self.HEADER_MAPPING.get(col_str, col_str)
            normalized.append(mapped)

        return normalized

    def _validate_required_columns(self, header: List[str], required: set, sheet_name: str):
        """필수 열 검증"""
        header_set = set(header)
        missing = required - header_set
        if missing:
            raise ValueError(f"{sheet_name} 시트에 필수 열이 없습니다: {missing}")

    def _parse_config_sheet(self, sheet) -> List[ConversationConfig]:
        """Configuration 시트 파싱"""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Configuration 시트가 비어있습니다.")

        header = self._normalize_header(rows[0])
        self._validate_required_columns(header, self.REQUIRED_CONFIG_COLS, "Configuration")

        configs = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):  # 빈 행 스킵
                continue

            row_dict = dict(zip(header, row))

            try:
                conversation_id = str(row_dict.get("conversation_id", "")).strip()
                context = str(row_dict.get("context", "")).strip()

                if not conversation_id or not context:
                    print(f"[Warning] Row {row_idx}: 필수 값 누락, 스킵")
                    continue

                config = ConversationConfig(
                    conversation_id=conversation_id,
                    context=context,
                    channel_name=str(row_dict.get("channel_name") or "general").strip(),
                    date=self._parse_date(row_dict.get("date")),
                    time_context=str(row_dict.get("time_context") or "").strip() or None,
                    message_count=int(row_dict.get("message_count") or 8),
                    style=str(row_dict.get("style") or "casual").strip().lower(),
                    language=str(row_dict.get("language") or "ko").strip().lower(),
                )
                configs.append(config)

            except Exception as e:
                print(f"[Warning] Row {row_idx} 파싱 실패: {e}")

        return configs

    def _parse_participants_sheet(self, sheet) -> List[Participant]:
        """Participants 시트 파싱"""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Participants 시트가 비어있습니다.")

        header = self._normalize_header(rows[0])
        self._validate_required_columns(
            header, self.REQUIRED_PARTICIPANT_COLS, "Participants"
        )

        participants = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):  # 빈 행 스킵
                continue

            row_dict = dict(zip(header, row))

            try:
                conversation_id = str(row_dict.get("conversation_id", "")).strip()
                participant_id = str(row_dict.get("participant_id", "")).strip()
                nickname = str(row_dict.get("nickname", "")).strip()
                role = str(row_dict.get("role", "other")).strip().lower()

                if not conversation_id or not participant_id or not nickname:
                    print(f"[Warning] Row {row_idx}: 필수 값 누락, 스킵")
                    continue

                participant = Participant(
                    conversation_id=conversation_id,
                    participant_id=participant_id,
                    nickname=nickname,
                    role=role,
                    avatar_seed=str(row_dict.get("avatar_seed") or "").strip() or None,
                    personality=str(row_dict.get("personality") or "").strip() or None,
                )
                participants.append(participant)

            except Exception as e:
                print(f"[Warning] Row {row_idx} 파싱 실패: {e}")

        return participants

    def _parse_date(self, date_val) -> Optional[datetime]:
        """날짜 파싱 (여러 형식 지원)"""
        if date_val is None:
            return None
        if isinstance(date_val, datetime):
            return date_val

        date_str = str(date_val).strip()
        formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y", "%d/%m/%Y"]
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    def _map_participants_to_configs(
        self, configs: List[ConversationConfig], participants: List[Participant]
    ):
        """참여자를 대화 설정에 매핑"""
        # conversation_id별로 참여자 그룹화
        participant_map = {}
        for p in participants:
            if p.conversation_id not in participant_map:
                participant_map[p.conversation_id] = []
            participant_map[p.conversation_id].append(p)

        for config in configs:
            config.participants = participant_map.get(config.conversation_id, [])

            # 참여자가 없으면 기본 참여자 생성
            if not config.participants:
                config.participants = [
                    Participant(config.conversation_id, "me", "나", "me"),
                    Participant(config.conversation_id, "other", "상대방", "other"),
                ]


# === Prompt Builder ===

class PromptBuilder:
    """Upstage API용 시스템/유저 프롬프트 생성기"""

    STYLE_DESCRIPTIONS = {
        "casual": "친근하고 자연스러운 일상 대화",
        "formal": "정중하고 격식 있는 대화",
        "dramatic": "극적이고 감정적인 대화",
    }

    LANGUAGE_NAMES = {
        "ko": "한국어",
        "en": "English",
        "ja": "日本語",
    }

    def build_system_prompt(self, config: ConversationConfig) -> str:
        """시스템 프롬프트 생성 (맥락 포함)"""
        participants_desc = self._build_participants_description(config.participants)
        style_desc = self.STYLE_DESCRIPTIONS.get(config.style, "자연스러운 대화")
        language_name = self.LANGUAGE_NAMES.get(config.language, "한국어")
        sender_ids = ", ".join([p.participant_id for p in config.participants])

        prompt = f"""당신은 Discord 스타일의 자연스러운 대화를 생성하는 AI입니다.
주어진 맥락에 맞는 {style_desc}를 JSON 형식으로 생성해주세요.

=== 대화 참여자 ===
{participants_desc}

=== 대화 규칙 ===
1. 반드시 JSON 형식으로만 응답하세요 (다른 텍스트 없이)
2. 정확히 {config.message_count}개의 메시지를 생성하세요
3. 언어: {language_name}
4. 대화자들이 자연스럽게 번갈아 대화하되, 엄격히 교대하지 않아도 됩니다
5. Discord 대화 특성:
   - 짧은 문장, 줄임말 사용 가능
   - 이모지 적절히 사용 (ㅋㅋ, ㅎㅎ, ㅠㅠ 등)
   - 비격식적 존댓말/반말 혼용
   - 자연스러운 대화 흐름 (질문-답변, 반응 등)
{self._build_time_context(config)}

=== 응답 JSON 형식 ===
{{"messages": [
  {{"sender": "참여자ID", "text": "메시지내용"}},
  ...
]}}

sender 필드는 반드시 다음 중 하나여야 합니다: {sender_ids}
"""
        return prompt

    def build_user_prompt(self, config: ConversationConfig) -> str:
        """유저 프롬프트 생성 (맥락 전달)"""
        return f"""다음 맥락으로 대화를 생성해주세요:

{config.context}

JSON 형식으로만 응답하세요."""

    def _build_participants_description(self, participants: List[Participant]) -> str:
        """참여자 설명 생성"""
        lines = []
        for p in participants:
            personality = f" - {p.personality}" if p.personality else ""
            role_label = f"[{p.role}]" if p.role in ["me", "other"] else ""
            lines.append(f"- {p.participant_id} ({p.nickname}) {role_label}{personality}")
        return "\n".join(lines)

    def _build_time_context(self, config: ConversationConfig) -> str:
        """시간 맥락 생성"""
        if config.time_context:
            return f"6. 대화 시간대: {config.time_context}"
        return ""


# === Upstage API Client ===

class UpstageClient:
    """Upstage Solar API 클라이언트"""

    def __init__(self, api_key: str):
        if not api_key:
            raise ValueError(
                "UPSTAGE_API_KEY 환경변수가 설정되지 않았습니다.\n"
                "사용법: UPSTAGE_API_KEY=up_xxx python scripts/generate_from_excel.py input.xlsx"
            )
        self.api_key = api_key
        self.base_url = UPSTAGE_BASE_URL

    async def generate_conversation(
        self, config: ConversationConfig, prompt_builder: PromptBuilder
    ) -> Optional[List[GeneratedMessage]]:
        """대화 생성 (재시도 로직 포함)"""
        system_prompt = prompt_builder.build_system_prompt(config)
        user_prompt = prompt_builder.build_user_prompt(config)

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }

        payload = {
            "model": DEFAULT_MODEL,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.8,
            "max_tokens": 2000,
        }

        for attempt in range(MAX_RETRIES):
            try:
                async with httpx.AsyncClient(timeout=60.0) as client:
                    response = await client.post(
                        f"{self.base_url}/chat/completions",
                        headers=headers,
                        json=payload,
                    )
                    response.raise_for_status()
                    result = response.json()

                    content = result["choices"][0]["message"]["content"]
                    messages = self._parse_response(content, config)
                    return messages

            except httpx.HTTPStatusError as e:
                print(
                    f"  [Retry {attempt + 1}/{MAX_RETRIES}] HTTP Error: {e.response.status_code}"
                )
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(RETRY_DELAY * (attempt + 1))
            except json.JSONDecodeError as e:
                print(f"  [Retry {attempt + 1}/{MAX_RETRIES}] JSON 파싱 실패: {e}")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(RETRY_DELAY * (attempt + 1))
            except Exception as e:
                print(f"  [Retry {attempt + 1}/{MAX_RETRIES}] Error: {e}")
                if attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(RETRY_DELAY * (attempt + 1))

        return None  # 모든 재시도 실패

    def _parse_response(
        self, content: str, config: ConversationConfig
    ) -> List[GeneratedMessage]:
        """API 응답 파싱"""
        # JSON 블록 추출
        try:
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                parts = content.split("```")
                if len(parts) >= 2:
                    content = parts[1]

            data = json.loads(content.strip())
            messages_data = data.get("messages", [])

        except json.JSONDecodeError:
            # JSON 배열만 있는 경우
            start = content.find("[")
            end = content.rfind("]") + 1
            if start != -1 and end > start:
                messages_data = json.loads(content[start:end])
            else:
                raise ValueError("JSON 파싱 실패")

        # 메시지 변환
        valid_senders = {p.participant_id for p in config.participants}
        messages = []

        for i, msg in enumerate(messages_data):
            sender = msg.get("sender", "other")
            if sender not in valid_senders:
                # 유효하지 않은 sender면 첫 번째 other 참여자로 매핑
                other_participants = [
                    p for p in config.participants if p.role != "me"
                ]
                if other_participants:
                    sender = other_participants[0].participant_id
                else:
                    sender = config.participants[0].participant_id

            text = msg.get("text", "")
            if text:
                messages.append(
                    GeneratedMessage(
                        id=i + 1,
                        sender=sender,
                        text=text,
                        time="",
                    )
                )

        return messages


# === Output Converter ===

class TalkStudioConverter:
    """TalkStudio JSON 형식 변환기"""

    def convert(
        self, config: ConversationConfig, messages: List[GeneratedMessage]
    ) -> dict:
        """TalkStudio 호환 JSON 생성"""
        base_time = self._get_base_time(config)

        # 프로필 매핑 생성
        profiles = self._build_profiles(config.participants)

        # 메시지 변환 (sender → role 매핑)
        converted_messages = []
        for i, msg in enumerate(messages):
            msg_time = base_time + timedelta(minutes=random.randint(0, 2) + i)

            # participant_id → role 매핑
            participant = next(
                (p for p in config.participants if p.participant_id == msg.sender),
                config.participants[0] if config.participants else None,
            )

            role = participant.role if participant else "other"

            converted_messages.append(
                {
                    "id": i + 1,
                    "sender": role,
                    "type": "text",
                    "text": msg.text,
                    "time": self._format_korean_time(msg_time),
                }
            )

        return {
            "config": {
                "theme": "discord",
                "capturedImage": None,
            },
            "statusBar": {
                "time": self._format_korean_time(base_time),
                "battery": random.randint(30, 95),
                "isWifi": True,
            },
            "profiles": profiles,
            "messages": converted_messages,
            "metadata": {
                "conversation_id": config.conversation_id,
                "channel_name": config.channel_name,
                "date": config.date.strftime("%Y-%m-%d") if config.date else None,
                "date_korean": self._format_korean_date(config.date),
                "context": config.context,
                "disclaimer": "AI 생성 샘플 데이터 - 실제 대화가 아님",
                "is_sample": True,
                "generated_at": datetime.now().isoformat(),
            },
        }

    def _get_base_time(self, config: ConversationConfig) -> datetime:
        """기준 시간 계산"""
        base_date = config.date or datetime.now()

        # time_context에서 시간대 추출 시도
        hour = random.randint(14, 22)  # 기본값
        if config.time_context:
            tc = config.time_context.lower()
            if "오전" in tc or "아침" in tc:
                hour = random.randint(8, 11)
            elif "점심" in tc:
                hour = random.randint(11, 13)
            elif "오후" in tc:
                hour = random.randint(14, 17)
            elif "저녁" in tc:
                hour = random.randint(18, 21)
            elif "밤" in tc or "늦은" in tc or "새벽" in tc:
                hour = random.randint(21, 23)

        return base_date.replace(
            hour=hour, minute=random.randint(0, 59), second=0, microsecond=0
        )

    def _build_profiles(self, participants: List[Participant]) -> dict:
        """프로필 빌드 (다중 참여자 지원)"""
        profiles = {}

        for p in participants:
            # role을 키로 사용 (me, other, 또는 participant_id)
            role_key = p.role if p.role in ["me", "other"] else p.participant_id
            profiles[role_key] = {
                "id": role_key,
                "name": p.nickname,
                "avatar": p.avatar_url,
            }

        return profiles

    def _format_korean_time(self, dt: datetime) -> str:
        """한국식 시간 포맷 (오후 3:30)"""
        period = "오전" if dt.hour < 12 else "오후"
        hour = dt.hour if dt.hour <= 12 else dt.hour - 12
        if hour == 0:
            hour = 12
        return f"{period} {hour}:{dt.minute:02d}"

    def _format_korean_date(self, dt: Optional[datetime]) -> Optional[str]:
        """한국식 날짜 포맷 (12월 22일)"""
        if not dt:
            return None
        return f"{dt.month}월 {dt.day}일"


# === Fallback Generator ===

class FallbackGenerator:
    """API 실패 시 사용할 기본 대화 생성기"""

    FALLBACK_TEMPLATES = {
        "ko": [
            {"sender": "other", "text": "안녕하세요!"},
            {"sender": "me", "text": "네, 안녕하세요~"},
            {"sender": "other", "text": "말씀하신 건에 대해서요"},
            {"sender": "me", "text": "네, 어떤 부분이 궁금하신가요?"},
            {"sender": "other", "text": "좀 더 자세히 알고 싶어서요"},
            {"sender": "me", "text": "아 네, 설명드릴게요!"},
            {"sender": "other", "text": "감사합니다 ㅎㅎ"},
            {"sender": "me", "text": "별말씀을요~"},
        ],
        "en": [
            {"sender": "other", "text": "Hey!"},
            {"sender": "me", "text": "Hi there~"},
            {"sender": "other", "text": "About what you mentioned..."},
            {"sender": "me", "text": "Sure, what would you like to know?"},
            {"sender": "other", "text": "I wanted more details"},
            {"sender": "me", "text": "Of course, let me explain!"},
            {"sender": "other", "text": "Thanks!"},
            {"sender": "me", "text": "No problem~"},
        ],
    }

    def generate(self, config: ConversationConfig) -> List[GeneratedMessage]:
        """Fallback 대화 생성"""
        template = self.FALLBACK_TEMPLATES.get(
            config.language, self.FALLBACK_TEMPLATES["ko"]
        )

        # 참여자 ID 매핑
        sender_map = {"me": "me", "other": "other"}

        if config.participants:
            me_participant = next(
                (p for p in config.participants if p.role == "me"), None
            )
            other_participant = next(
                (p for p in config.participants if p.role != "me"), None
            )

            if me_participant:
                sender_map["me"] = me_participant.participant_id
            if other_participant:
                sender_map["other"] = other_participant.participant_id

        messages = []
        for i, msg in enumerate(template[: config.message_count]):
            messages.append(
                GeneratedMessage(
                    id=i + 1,
                    sender=sender_map.get(msg["sender"], msg["sender"]),
                    text=msg["text"],
                    time="",
                )
            )

        return messages


# === Main Entry Point ===

async def main():
    """메인 실행 함수"""
    parser = argparse.ArgumentParser(
        description="Excel에서 Discord 대화 생성",
        epilog="Example: UPSTAGE_API_KEY=up_xxx python %(prog)s input.xlsx",
    )
    parser.add_argument("input", help="입력 엑셀 파일 경로")
    parser.add_argument(
        "--output",
        "-o",
        default="scripts/generated_chats.json",
        help="출력 JSON 파일 경로 (기본: scripts/generated_chats.json)",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("TalkStudio - Excel 기반 Discord 대화 생성기")
    print("=" * 60)

    # 1. 엑셀 파싱
    print(f"\n[1/3] 엑셀 파일 읽는 중: {args.input}")
    try:
        excel_parser = ExcelParser(args.input)
        configs = excel_parser.parse()
        print(f"  -> {len(configs)}개 대화 설정 로드 완료")

        for i, c in enumerate(configs):
            print(f"     {i + 1}. {c.conversation_id}: {len(c.participants)}명 참여")
    except Exception as e:
        print(f"  [Error] 엑셀 파싱 실패: {e}")
        sys.exit(1)

    # 2. API 클라이언트 초기화
    print("\n[2/3] 대화 생성 중...")
    try:
        client = UpstageClient(UPSTAGE_API_KEY)
    except ValueError as e:
        print(f"  [Error] {e}")
        sys.exit(1)

    prompt_builder = PromptBuilder()
    converter = TalkStudioConverter()
    fallback = FallbackGenerator()

    all_outputs = []

    for i, config in enumerate(configs):
        print(f"\n  [{i + 1}/{len(configs)}] {config.conversation_id}")
        print(f"    맥락: {config.context[:50]}{'...' if len(config.context) > 50 else ''}")
        print(f"    참여자: {', '.join([p.nickname for p in config.participants])}")

        # API 호출
        messages = await client.generate_conversation(config, prompt_builder)

        if messages is None:
            print("    -> API 실패, Fallback 사용")
            messages = fallback.generate(config)
        else:
            print(f"    -> 성공! {len(messages)}개 메시지 생성")

        # TalkStudio 형식으로 변환
        output = converter.convert(config, messages)
        all_outputs.append(output)

    # 3. JSON 저장
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"\n[3/3] 결과 저장 중: {args.output}")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_outputs, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 60)
    print(f"완료! {len(all_outputs)}개 대화 생성됨")
    print(f"출력 파일: {args.output}")
    print("=" * 60)

    # 요약 출력
    print("\n생성된 대화 요약:")
    for i, output in enumerate(all_outputs):
        meta = output["metadata"]
        msg_count = len(output["messages"])
        print(
            f"  {i + 1}. {meta['conversation_id']} - {meta['channel_name']} ({msg_count}개 메시지)"
        )

    print("\n다음 단계:")
    print("  이미지 캡처: node scripts/capture_discord_chats.cjs")

    return all_outputs


if __name__ == "__main__":
    asyncio.run(main())
