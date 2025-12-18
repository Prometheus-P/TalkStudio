"""Excel parsing service."""

import io
import uuid
from datetime import datetime, timezone

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.schemas.chat import ChatMessage, MessageType, SpeakerType


class ExcelService:
    """Service for parsing Excel files into chat messages."""

    # Expected column headers (case-insensitive)
    COLUMN_MAPPINGS = {
        "speaker": ["speaker", "발신자", "보낸사람", "who", "from"],
        "text": ["text", "내용", "메시지", "message", "content"],
        "type": ["type", "타입", "유형"],
        "time": ["time", "시간", "timestamp", "날짜"],
        "name": ["name", "이름", "닉네임", "nickname"],
    }

    async def parse_excel(
        self,
        file_content: bytes,
        filename: str,
    ) -> tuple[list[ChatMessage], dict]:
        """
        Parse Excel file content into chat messages.

        Returns:
            Tuple of (messages, stats)
        """
        try:
            workbook = load_workbook(
                filename=io.BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except InvalidFileException:
            raise ValueError("Invalid Excel file format")
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {e}")

        # Use first sheet by default
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Excel file has no sheets")

        # Get all rows
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")

        # Find header row and map columns
        header_row_idx, column_map = self._find_header_row(rows)
        if column_map is None:
            raise ValueError(
                "Could not find required columns. "
                "Please include 'speaker' and 'text' columns."
            )

        # Parse data rows
        messages = []
        errors = []
        warnings = []

        data_rows = rows[header_row_idx + 1 :]
        total_rows = len(data_rows)

        for row_idx, row in enumerate(data_rows, start=header_row_idx + 2):
            try:
                message = self._parse_row(row, column_map, row_idx)
                if message:
                    messages.append(message)
            except ValueError as e:
                errors.append({"row": row_idx, "error": str(e)})
            except Exception as e:
                errors.append({"row": row_idx, "error": f"Unexpected error: {e}"})

        if not messages:
            raise ValueError("No valid messages found in Excel file")

        # Add warnings
        if errors:
            warnings.append(f"{len(errors)} rows had errors and were skipped")

        stats = {
            "total_rows": total_rows,
            "parsed_rows": len(messages),
            "errors": errors[:10],  # Limit errors returned
            "warnings": warnings,
        }

        workbook.close()
        return messages, stats

    def _find_header_row(
        self,
        rows: list[tuple],
        max_search: int = 10,
    ) -> tuple[int, dict | None]:
        """
        Find the header row and create column mapping.

        Returns:
            Tuple of (header_row_index, column_mapping)
        """
        for row_idx, row in enumerate(rows[:max_search]):
            column_map = self._try_map_columns(row)
            if column_map and "speaker" in column_map and "text" in column_map:
                return row_idx, column_map

        return -1, None

    def _try_map_columns(self, row: tuple) -> dict[str, int] | None:
        """Try to map header row to column indices."""
        if not row:
            return None

        column_map = {}

        for col_idx, cell in enumerate(row):
            if cell is None:
                continue

            cell_str = str(cell).strip().lower()

            for field, aliases in self.COLUMN_MAPPINGS.items():
                if cell_str in aliases:
                    column_map[field] = col_idx
                    break

        return column_map if column_map else None

    def _parse_row(
        self,
        row: tuple,
        column_map: dict[str, int],
        row_idx: int,
    ) -> ChatMessage | None:
        """Parse a single row into a ChatMessage."""
        if not row or all(cell is None for cell in row):
            return None

        # Get speaker (required)
        speaker_idx = column_map.get("speaker")
        if speaker_idx is None or speaker_idx >= len(row):
            raise ValueError("Missing speaker column")

        speaker_raw = row[speaker_idx]
        if speaker_raw is None:
            return None  # Skip empty rows

        speaker_str = str(speaker_raw).strip().lower()
        if speaker_str in ("me", "나", "본인", "1"):
            speaker = SpeakerType.ME
        elif speaker_str in ("system", "시스템"):
            speaker = SpeakerType.SYSTEM
        else:
            speaker = SpeakerType.OTHER

        # Get text (required)
        text_idx = column_map.get("text")
        if text_idx is None or text_idx >= len(row):
            raise ValueError("Missing text column")

        text_raw = row[text_idx]
        if text_raw is None:
            return None

        text = str(text_raw).strip()
        if not text:
            return None

        # Get type (optional)
        msg_type = MessageType.TEXT
        type_idx = column_map.get("type")
        if type_idx is not None and type_idx < len(row) and row[type_idx]:
            type_str = str(row[type_idx]).strip().lower()
            if type_str in ("emoji", "이모지", "이모티콘"):
                msg_type = MessageType.EMOJI
            elif type_str in ("image", "이미지", "사진"):
                msg_type = MessageType.IMAGE
            elif type_str in ("system", "시스템"):
                msg_type = MessageType.SYSTEM

        # Get name (optional)
        name = None
        name_idx = column_map.get("name")
        if name_idx is not None and name_idx < len(row) and row[name_idx]:
            name = str(row[name_idx]).strip()

        # Get time (optional)
        timestamp = datetime.now(timezone.utc)
        time_idx = column_map.get("time")
        if time_idx is not None and time_idx < len(row) and row[time_idx]:
            time_val = row[time_idx]
            if isinstance(time_val, datetime):
                timestamp = time_val.replace(tzinfo=timezone.utc)

        return ChatMessage(
            id=str(uuid.uuid4()),
            speaker=speaker,
            speaker_name=name,
            text=text,
            type=msg_type,
            timestamp=timestamp,
        )
